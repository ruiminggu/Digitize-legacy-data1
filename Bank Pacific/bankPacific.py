import re
import os
import string
import shutil
from convert_to_txt import split_pdf, convert_folder_images, convert_folder_pdf_to_png
from openpyxl import Workbook

file_path = 'BankofPacific/png_files/txt_files'
output_path = 'BankofPacific/png_files/txt_files/result.csv'

# Create an Excel workbook
workbook = Workbook()
workbook2 = Workbook()
sheet = workbook.active
sheet2 = workbook2.active

def generate_ctr_id(trans_date, ctr_count):
    date_parts = trans_date.split('/')
    formatted_date = f"{date_parts[0]:0>2}{date_parts[1]:0>2}{date_parts[2]}"
    return f"{formatted_date}{ctr_count:03d}"

def read_CTR_file(file_path, sheet, ctr_tracker, ctrid_list):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
        lines = content.splitlines()

        trans_date = ""
        branches_id = "Koror Branch"
        transaction_type = ""
        amount = ""
        fullNameOfFinancialInstitution = "BANKPACIFIC"
        typeOfFinancialInstitution = "Depository Institution"
        cash_in_amount = ""
        cash_out_amount = ""

        for text in lines:
            trans_date_match = re.search(r"Date of Transaction: (\d{1,2}/\d{1,2}/\d{4})", text)
            if trans_date_match:
                trans_date = trans_date_match.group(1)
                if trans_date in ctr_tracker:
                    ctr_tracker[trans_date] += 1
                else:
                    ctr_tracker[trans_date] = 1
                ctr_id = generate_ctr_id(trans_date, ctr_tracker[trans_date])
                ctrid_list.append(ctr_id)

            deposit_match = re.search(r"Total\s*Cash\s*In\s*\$\s*([\d,]+(?:\.\d+)?)", text)
            withdrawal_match = re.search(r"Total\s*Cash\s*Out\s*\$\s*([\d,]+(?:\.\d+)?)", text)

            if deposit_match:
                cash_in_amount = deposit_match.group(1).replace(',', '').replace(' ', '')
                if float(cash_in_amount) > 0:
                    transaction_type = "deposit"
                    amount = cash_in_amount

            if withdrawal_match:
                cash_out_amount = withdrawal_match.group(1).replace(',', '').replace(' ', '')
                if float(cash_out_amount) > 0 and (transaction_type != "deposit" or float(cash_in_amount) == 0):
                    transaction_type = "withdrawal"
                    amount = cash_out_amount

        if trans_date:
            if transaction_type == "deposit":
                amount = cash_in_amount
            elif transaction_type == "withdrawal":
                amount = cash_out_amount
            sheet.append([ctr_id, trans_date, branches_id, transaction_type, amount, fullNameOfFinancialInstitution, typeOfFinancialInstitution])

def read_folder_CTR(folder_path, output_path, ctrid_list):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["CTRID","dateOfTransaction","nameOfBranchOfficeAgency","cashDirection","cashAmount","fullNameOfFinancialInstitution","typeOfFinancialInstitution"])
    ctr_tracker = {}

    filenames = sorted([f for f in os.listdir(folder_path) if f.endswith('.txt')])

    for filename in filenames:
        file_path = os.path.join(folder_path, filename)
        read_CTR_file(file_path, sheet, ctr_tracker, ctrid_list)

    workbook.save(output_path)

def read_PIT_file(file_path, sheet2, ctrid_list, ctrid_index):
    with open(file_path) as file:
        lines = file.readlines()

    if ctrid_index < len(ctrid_list):
        ctrid = ctrid_list[ctrid_index]
    else:
        ctrid = ""

    print(f"CTR index at start: {ctrid_index}")
    print(f"Initial CTR ID: {ctrid}")  # Debug print

    lastNameOrNameOfEntity = ""
    firstName = ""
    middleName = ""
    gender = ""
    occupation = ""
    address = ""
    city = ""
    state = ""
    zipcode = ""
    address_country = ""
    dob = ""
    contact_number = ""
    id_type = ""
    id_number = ""
    id_country = ""
    account_number = ""
    email = ""
    relationship = ""
    cash_direction = ""
    cash_amount = ""

    for line in lines:
        line = line.strip()

        if "End of Report" in line:
            print(f"Encountered 'End of Report' in line: {line}")
            ctrid_index += 1
            if ctrid_index < len(ctrid_list):
                ctrid = ctrid_list[ctrid_index]
                print(f"Switching to next CTR ID: {ctrid}")  # Debug print
            else:
                print("No more CTR IDs available")  # Debug print
                ctrid = ""

        # print(f"Current CTR index: {ctrid_index}")

        last_name_match = re.search(r"4[,.] Last Name/Entity Name: (.*?)$", line)
        if last_name_match:
            lastNameOrNameOfEntity = last_name_match.group(1).strip()

        first_name_match = re.search(r"5[,.] First Name: (.*?)( 6[,.]|$)", line)
        if first_name_match:
            firstName = first_name_match.group(1).strip()

        middle_name_match = re.search(r"6[,.] Middle Name: (.*?)( |$)", line)
        if middle_name_match:
            middleName = middle_name_match.group(1).strip()

        gender_match = re.search(r"7[,.] Gender: (\w+)", line)
        if gender_match:
            gender = gender_match.group(1)

        occupation_match = re.search(r"9[,.] Occupation or Type of Business: (.*?)$", line)
        if occupation_match:
            occupation = occupation_match.group(1).strip()

        address_match = re.search(r"10[,.] Address: (.*?)$", line)
        if address_match:
            address = address_match.group(1).strip()

        city_match = re.search(r"11[,.] City: (.*?) 12[,.] State:", line)
        if city_match:
            city = city_match.group(1).strip()

        state_match = re.search(r"12[,.] State: (.*?) 13[,.] Zip Code:", line)
        if state_match:
            state = state_match.group(1).strip()

        zipcode_match = re.search(r"13[,.] Zip Code: (\d+)", line)
        if zipcode_match:
            zipcode = zipcode_match.group(1).strip()

        country_match = re.search(r"14[,.] Country: (.*?) 9a[,.]", line)
        if country_match:
            address_country = country_match.group(1).strip()

        dob_match = re.search(r"17[,.] Date of Birth: (.*?) 7[,.]", line)
        if dob_match:
            dob = dob_match.group(1).strip()

        contact_number_match = re.search(r"18[,.] Contact Phone Number[:;] (\d+)", line)
        if contact_number_match:
            contact_number = contact_number_match.group(1).strip()

        email_match = re.search(r"19[,.] Email Address: (.*?)$", line)
        if email_match:
            email = email_match.group(1).strip()

        id_type_match = re.search(r"20[,.] Type of Identification: (.*?) Other", line)
        if id_type_match:
            id_type = id_type_match.group(1).strip()

        id_number_match = re.search(r"Number: (.*?) Country", line)
        if id_number_match:
            id_number = id_number_match.group(1).strip()

        id_country_match = re.search(r"Country: (.*?) State", line)
        if id_country_match:
            id_country = id_country_match.group(1).strip()

        account_number_match = re.search(r"Account Number\(\w*\): (\d+)", line)
        if account_number_match:
            account_number = account_number_match.group(1).strip()

        relationship_match = re.search(r"2[,.] Person Involved in Transaction: (.*?)$", line)
        if relationship_match:
            relationship = relationship_match.group(1).strip()

        cash_in_match = re.search(r"21[,.] Cash In Amount for Individual or Entity: \$ ([\d,]+)", line)
        if cash_in_match:
            cash_direction = "deposit"
            cash_amount = cash_in_match.group(1).replace(",", "").strip()

        cash_out_match = re.search(r"22[,.] Cash Out Amount for Individual or Entity: \$ ([\d,]+)", line)
        if cash_out_match:
            cash_out_amount = cash_out_match.group(1).replace(",", "").strip()
            if float(cash_out_amount) > 0:  # Only update if there is a non-zero cash out amount
                cash_direction = "withdrawal"
                cash_amount = cash_out_amount

        begin = re.search(r"PART I ", line)
        if (lastNameOrNameOfEntity and begin) or lines[-1].strip() == line:
            if any([lastNameOrNameOfEntity, firstName, middleName, gender, occupation, address, city, state, zipcode,
                    address_country, dob, contact_number, id_type, id_number, id_country, account_number, email,
                    relationship, cash_direction, cash_amount]):
                # Debug print to verify the information being appended
                print(f"Appending row with CTR ID: {ctrid}, Last Name: {lastNameOrNameOfEntity}")  # Debug print
                # Write the extracted information to the Excel sheet
                sheet2.append([
                    ctrid, lastNameOrNameOfEntity, firstName, middleName, gender, occupation, address, city, state,
                    zipcode,
                    address_country, dob, contact_number, id_type, id_number, id_country, account_number, email,
                    relationship, cash_direction, cash_amount
                ])
            lastNameOrNameOfEntity = ""
            firstName = ""
            middleName = ""
            occupation = ""
            address = ""
            city = ""
            state = ""
            zipcode = ""
            address_country = ""
            dob = ""
            contact_number = ""
            id_type = ""
            id_number = ""
            id_country = ""
            account_number = ""
            email = ""
            cash_direction = ""
            cash_amount = ""

    return ctrid_index  # Return the updated index

def read_folder_PIT(folder_path, output_path, ctrid_list):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "CTRID", "lastNameOrNameOfEntity", "firstName", "middleName", "gender", "occupationOrTypeOfBusiness",
        "address", "addressCity", "addressState", "zipCode", "addressCountry", "dateOfBirth",
        "contactPhoneNumber", "idType", "idNumber", "idCountry", "accountNumbers", "emailAddress",
        "relationshipToTransaction", "cashDirection", "cashAmount"
    ])
    ctrid_index = 0

    filenames = sorted([f for f in os.listdir(folder_path) if f.endswith('.txt')])

    for filename in filenames:
        file_path = os.path.join(folder_path, filename)
        ctrid_index = read_PIT_file(file_path, sheet, ctrid_list, ctrid_index)
        print(f"CTR index after file: {ctrid_index}")

    workbook.save(output_path)

def reset_folders(folder_path):
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)
    os.makedirs(folder_path)
    os.makedirs(os.path.join(folder_path, 'png_files'))
    os.makedirs(os.path.join(folder_path, 'png_files/txt_files'))

if __name__ == '__main__':
    folder_path = "rmtest"
    pdf_file = "BankPacific.pdf"
    # reset_folders(folder_path)
    # split_pdf(folder_path, pdf_file)
    # convert_folder_images(convert_folder_pdf_to_png(folder_path))

    ctrid_list = []
    read_folder_CTR(folder_path + "/png_files/txt_files", "rmctr.xlsx", ctrid_list)
    read_folder_PIT(folder_path + "/png_files/txt_files", "rmpit.xlsx", ctrid_list)
    print(ctrid_list)
