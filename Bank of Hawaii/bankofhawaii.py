import re
import os
import openpyxl

from convert_to_txt import split_pdf, convert_folder_images, convert_folder_pdf_to_png

file_path = 'PDFs/png_files/txt_files'
output_path = 'PDFs/png_files/txt_files/result.csv'
from openpyxl import Workbook

# Create an Excel workbook
workbook = Workbook()
workbook2 = Workbook()
sheet = workbook.active
sheet2 = workbook2.active

def read_file(file_path, sheet):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
        lines = content.splitlines()

        ctr_id = ""
        trans_date = ""
        branches_id = ""
        transaction_type = ""
        amount = ""
        fullNameOfFinancialInstitution = "Bank of Hawaii"
        typeOfFinancialInstitution = "Depository Institution"
        cash_in_amount = ""
        cash_out_amount = ""

        for text in lines:
            # Extract CTR ID
            ctr_id_match = re.search(r"CTR (\d+)", text)
            if ctr_id_match:
                ctr_id = ctr_id_match.group(1)

            # Extract TRANS DATE
            trans_date_match = re.search(r"TRANS DATE (\d{2}/\d{2}/\d{4})", text)
            if trans_date_match:
                trans_date = trans_date_match.group(1)

            # Extract BRANCHES ID
            branches_id_match = re.search(r"BRANCHES (\d+)", text)
            if branches_id_match and branches_id_match.group(1) == "0000037":
                branches_id = "Koror Branch"

            # Extract Deposit or Withdrawals
            deposit_match = re.search(r"TOTAL\s*CASH[-~\s]*IN\s*\$(\d+(?:,\s*\d+)*(?:\.\s*\d+)?)", text)
            withdrawal_match = re.search(r"TOTAL\s*CASH[-~\s]*OUT\s*\$(\d+(?:,\s*\d+)*(?:\.\s*\d+)?)", text)

            if deposit_match:
                transaction_type = "deposit"
                cash_in_amount = deposit_match.group(1).replace(',', '').replace(' ', '')
            if withdrawal_match:
                transaction_type = "withdrawal"
                cash_out_amount = withdrawal_match.group(1).replace(',', '').replace(' ', '')

        # Add to sheet if CTR ID is present
        if ctr_id:
            if transaction_type == "deposit":
                amount = cash_in_amount
            elif transaction_type == "withdrawal":
                amount = cash_out_amount
            sheet.append([ctr_id, trans_date, branches_id, transaction_type, amount, fullNameOfFinancialInstitution, typeOfFinancialInstitution])

# Example usage with openpyxl to create a workbook and sheet
def read_folder_CTR(folder_path,output_path):
    workbook = Workbook()
    sheet = workbook.active
    folder_path = folder_path + "/png_files/txt_files"
    sheet.append(["CTRID","dateOfTransaction","nameOfBranchOfficeAgency","cashDirection","cashAmount","fullNameOfFinancialInstitution","typeOfFinancialInstitution"])
    for filename in os.listdir(folder_path):
        txt_file_path = os.path.join(folder_path, filename)
        read_file(txt_file_path,sheet)
    workbook.save(output_path)

def read_file_PIT(file_path, sheet2):
    with open(file_path, "r", encoding='utf-8') as file:
        lines = file.readlines()

    # Initialize variables to store extracted information
    ctrid = ""
    lastNameOrNameOfEntity = ""
    firstName = ""
    middleName = ""
    gender = ""
    occupation = ""
    address = ""
    city = ""
    state = ""
    zipcode = ""
    address_country = ""
    dob = ""
    contact_number = ""
    id_type = ""
    id_number = ""
    id_country = ""
    account_number = ""
    cash_direction = ""
    cash_amount = ""
    email = ""
    relationship = ""
    is_person = False  # Default to False; will be updated based on DOB presence

    def append_to_sheet():
        if lastNameOrNameOfEntity or firstName or middleName:
            sheet2.append([ctrid, lastNameOrNameOfEntity, firstName, middleName, gender, occupation, address, city, state, zipcode, address_country, dob,
                           contact_number, id_type, id_number, id_country, account_number, cash_direction,
                           cash_amount, email, relationship])

    for line in lines:
        line = line.strip()

        # Check if the line indicates the start of a new person's information
        relationship_match = re.search(r"COND/BENEF ([BCS])", line)
        if relationship_match:
            # If there's already a name, append the previous person's data to the sheet
            append_to_sheet()

            # Extract and set the relationship for the new person
            relationship_code = relationship_match.group(1)
            if relationship_code == "B":
                relationship = "Person on whose behalf transaction was conducted"
            elif relationship_code == "C":
                relationship = "Person conducting transaction for another"
            elif relationship_code == "S":
                relationship = "Person conducting transaction on own behalf"

            # Reset variables for the new person, except ctrid
            lastNameOrNameOfEntity = ""
            firstName = ""
            middleName = ""
            gender = ""
            occupation = ""
            address = ""
            city = ""
            state = ""
            zipcode = ""
            address_country = ""
            dob = ""
            contact_number = ""
            id_type = ""
            id_number = ""
            id_country = ""
            account_number = ""
            cash_direction = ""
            cash_amount = ""
            email = ""
            is_person = False

        # Extract CTRID
        ctrid_match = re.search(r"CTR (\d+)", line)
        if ctrid_match:
            ctrid = ctrid_match.group(1)

        # Extract Date of Birth
        dob_match = re.search(r"BIRTH DATE (\d+/\d+/\d+)", line)
        if dob_match:
            dob = dob_match.group(1)
            is_person = True  # Presence of DOB indicates a person

        # Extract Name
        name_match = re.search(r"NAME (.+?) ALT NAME", line)
        if name_match:
            full_name = name_match.group(1)
            if is_person:
                name_parts = full_name.split()
                firstName = name_parts[0]
                if len(name_parts) == 2:
                    middleName = ""
                    lastNameOrNameOfEntity = name_parts[1]
                elif len(name_parts) > 2:
                    middleName = " ".join(name_parts[1:-1])
                    lastNameOrNameOfEntity = name_parts[-1]
            else:
                lastNameOrNameOfEntity = full_name
                firstName = ""
                middleName = ""

        # Extract Occupation
        occupation_match = re.search(r"OCCUP (.+)", line)
        if occupation_match:
            occupation_match1 = re.search(r"OCCUP (.+?) ENTITY", line)
            if occupation_match1:
                occupation = occupation_match1.group(1)
            else:
                occupation = occupation_match.group(1)

        # Extract Address State
        state_match = re.search(r"STATE (.+?) ZIP", line)
        if state_match:
            state = state_match.group(1)

        # Extract Address
        address_match = re.search(r"ADDR (.+) ", line)
        city_match = re.search(r"CITY (.+?) STATE", line)
        if address_match and not city_match:
            address = address_match.group(1)

        # Extract Address City
        if address_match and city_match:
            address_match = re.search(r"ADDR (.+?) CITY", line)
            address = address_match.group(1)
            city = city_match.group(1)

        # Extract Email
        email_match = re.search(r"EMAIL (.+?) BIRTH", line)
        if email_match:
            email = email_match.group(1)

        # Extract Zipcode
        zipcode_match = re.search(r"ZIP (\d+)", line)
        if zipcode_match:
            zipcode = zipcode_match.group(1)

        # Extract Address Country
        country_match = re.search(r"CNTRY (.+?) ", line)
        if country_match:
            address_country = country_match.group(1)

        # Extract Contact Number
        contact_number_match = re.search(r"PHONE (\d+)", line)
        if contact_number_match:
            contact_number = contact_number_match.group(1)

        # Extract ID Type, ID Number, and ID Country
        id_match = re.search(r"ID METHOD (.+?) ID", line)
        if id_match:
            id_type = id_match.group(1)
        id_type_match = re.search(r"ID # (.+) ID", line)
        if id_type_match:
            id_number = id_type_match.group(1)
        id_country_match = re.search(r"CNTRY (.+?) ID", line)
        if id_country_match:
            id_country = id_country_match.group(1)

        # Double-check the ID type
        id_match_type_des_ = re.search(r"ID OTHER DESC (.+)", line)
        if (id_type == "" or id_type == "OTHER") and id_match_type_des_:
            id_type = id_match_type_des_.group(1)

        # Extract Account Number
        account_number_match = re.search(r"ACCOUNT NUMBER\(S\) (.+)", line)
        if account_number_match:
            account_number = account_number_match.group(1).split()
            account_number = ', '.join(account_number)

        # Extract Cash Direction and Cash Amount
        cash_in_match = re.search(r"CASH-IN AMT \$(\d+(?:,\s*\d+)*(?:\.\s*\d+)?)", line)
        if cash_in_match:
            cash_amount = "$" + cash_in_match.group(1)
            cash_direction = "Deposit"
        # Extract misplaced cash amount
        cash_amount_check = re.search(r"^\$([\d,]+)$", line)
        if cash_amount_check:
            cash_amount = "$" + cash_amount_check.group(1)
        # Extract Cash-Out amount and type
        cash_out_match = re.search(r"CASH-OUT AMT \$(\d+(?:,\s*\d+)*(?:\.\s*\d+)?)", line)
        if cash_out_match:
            cash_amount = "$" + cash_out_match.group(1)
            cash_direction = "Withdrawal"
        # Extract misplaced cash account
        account_number_check = re.search(r"^(\d{8})$", line)
        if account_number_check:
            account_number = account_number_check.group(1)

        # Extract Gender
        gender_match = re.search(r"GENDER ([AB])", line)
        if gender_match:
            gender_code = gender_match.group(1)
            if gender_code == "A":
                gender = "Male"
            elif gender_code == "B":
                gender = "Female"

    # Append the last person's data to the sheet
    append_to_sheet()

def read_folder_PIT(folder_path, output_path):
    workbook = Workbook()
    sheet = workbook.active
    folder_path = folder_path + "/png_files/txt_files"
    sheet.append(["CTRID", "lastNameOrNameOfEntity", "firstName", "middleName", "gender", "occupationOrTypeOfBusiness", "address", "addressCity", "addressState", "zipCode", "addressCountry", "dateOfBirth",
                  "contactPhoneNumber", "idType", "idNumber", "idCountry", "accountNumbers", "cashDirection",
                  "cashAmount", "emailAddress", "relationshipToTransaction"])
    for filename in os.listdir(folder_path):
        txt_file_path = os.path.join(folder_path, filename)
        read_file_PIT(txt_file_path, sheet)
    workbook.save(output_path)



if __name__ == '__main__':
    # the first should enter the name of the destination folder, the second is the pdf name
    split_pdf("rmtest", "Bank_of_Hawaii.pdf")
    # enter the first again.
    convert_folder_images(convert_folder_pdf_to_png("rmtest"))

    # enter the first again, secod input should be the name of PIT report
    read_folder_PIT("rmtest", "rmpit.xlsx")
    # enter the first again, secod input should be the name of CTR report
    read_folder_CTR("rmtest", "rmctr.xlsx")
